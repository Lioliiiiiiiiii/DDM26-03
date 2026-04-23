#!/usr/bin/env python3
"""Build normalized technology page data from the source Excel workbooks."""

from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from statistics import mean
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


@dataclass(frozen=True)
class TechnologySpec:
    key: str
    slug: str
    name: str
    short_name: str
    color: str
    aliases: tuple[str, ...]


TECH_SPECS: tuple[TechnologySpec, ...] = (
    TechnologySpec(
        key="descriptive-ai",
        slug="descriptive-ai",
        name="Descriptive AI",
        short_name="Descriptive AI",
        color="#f59e0b",
        aliases=("descriptive ai",),
    ),
    TechnologySpec(
        key="agentic-generative-ai",
        slug="agentic-gen-ai",
        name="Agentic & GenAI",
        short_name="Agentic & GenAI",
        color="#FF5EA8",
        aliases=("agentic gen ai", "agenticgenai", "generative ai", "agentic ai"),
    ),
    TechnologySpec(
        key="blockchain-decentralized-systems",
        slug="blockchain-decentralized-systems",
        name="Blockchain & Decentralized Systems",
        short_name="Blockchain & Decentralized Systems",
        color="#FF7A45",
        aliases=("blockchain", "blockchain decentralized systems"),
    ),
    TechnologySpec(
        key="robotics-automation",
        slug="robotics-automation",
        name="Robotics & Automation",
        short_name="Robotics & Automation",
        color="#34D1BF",
        aliases=("robotics", "robotics automation"),
    ),
    TechnologySpec(
        key="quantum-computing",
        slug="quantum-computing",
        name="Quantum Computing",
        short_name="Quantum Computing",
        color="#6EA8FF",
        aliases=("quantum", "quantum computing"),
    ),
)

TECH_BY_KEY = {item.key: item for item in TECH_SPECS}

INDUSTRY_LOOKUP = {
    "energy": ("energy", "Energy"),
    "materials": ("materials", "Materials"),
    "industrials": ("industrials", "Industrials"),
    "consumergoods": ("consumer-goods", "Consumer Goods"),
    "healthcare": ("health-care", "Health Care"),
    "financialsservices": ("financial-services", "Financial Services"),
    "financialservices": ("financial-services", "Financial Services"),
    "informationtechnology": ("information-technology", "Information Technology"),
    "communicationservicescreative": (
        "communication-creative-services",
        "Communication & Creative Services",
    ),
    "realestate": ("real-estate", "Real Estate"),
    "automotivetransport": ("automotive-transport", "Automotive & Transport"),
    "it": ("information-technology", "Information Technology"),
    "financials": ("financial-services", "Financial Services"),
    "financial": ("financial-services", "Financial Services"),
    "automotive": ("automotive-transport", "Automotive & Transport"),
}

RESEARCH_POINT_YEARS = (2020, 2021, 2022, 2023, 2024, 2025)

Q9_INDUSTRY_BLOCKS = (
    ("Energy", "U", "AI"),
    ("Industrials", "AJ", "AX"),
    ("Materials", "AY", "BM"),
    ("Consumer Goods", "BN", "CB"),
    ("Health Care", "CC", "CQ"),
    ("Financial Services", "CR", "DF"),
    ("Information Technology", "DG", "DU"),
    ("Real Estate", "DV", "EJ"),
    ("Communication", "EK", "EY"),
    ("Automotive & Transport", "EZ", "FN"),
)

TECH_TO_USECASE_CHUNK_INDEX = {
    "descriptive-ai": 0,
    "agentic-generative-ai": 1,
    "blockchain-decentralized-systems": 2,
    "robotics-automation": 3,
    "quantum-computing": 4,
}


def normalize_text(value: str | None) -> str:
    if not value:
        return ""
    return re.sub(r"[^a-z0-9]+", "", value.lower())


def parse_number(value: Any) -> float:
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        trimmed = value.strip()
        if not trimmed:
            return 0.0
        try:
            return float(trimmed.replace(",", ""))
        except ValueError:
            return 0.0
    return 0.0


def parse_year(value: Any) -> int | None:
    if value is None:
        return None
    if isinstance(value, datetime):
        return value.year
    if isinstance(value, date):
        return value.year
    if isinstance(value, (int, float)):
        year = int(value)
        return year if 1900 <= year <= 2100 else None
    return None


def format_founded_date(value: Any, precision: Any) -> str:
    normalized_precision = normalize_text(precision if isinstance(precision, str) else "")
    if isinstance(value, datetime):
        if "day" in normalized_precision:
            return value.strftime("%Y-%m-%d")
        if "month" in normalized_precision:
            return value.strftime("%b %Y")
        return str(value.year)
    if isinstance(value, date):
        return str(value.year)
    if isinstance(value, (int, float)):
        return str(int(value))
    if isinstance(value, str):
        stripped = value.strip()
        if stripped:
            return stripped
    return "n/a"


def clean_label(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    return re.sub(r"\s+", " ", value).strip()


def get_tech_key(raw_value: Any) -> str | None:
    if not isinstance(raw_value, str):
        return None
    candidate = normalize_text(raw_value)
    if not candidate:
        return None

    for spec in TECH_SPECS:
        alias_candidates = (spec.name, spec.short_name, *spec.aliases)
        for alias in alias_candidates:
            alias_norm = normalize_text(alias)
            if alias_norm and (alias_norm in candidate or candidate in alias_norm):
                return spec.key

    return None


def tech_keys_in_text(raw_value: Any) -> list[str]:
    if not isinstance(raw_value, str):
        return []
    candidate = normalize_text(raw_value)
    if not candidate:
        return []

    keys: list[str] = []
    for spec in TECH_SPECS:
        alias_candidates = (spec.name, spec.short_name, *spec.aliases)
        if any(normalize_text(alias) in candidate for alias in alias_candidates if alias):
            keys.append(spec.key)
    return keys


def get_industry(raw_value: Any) -> tuple[str, str] | None:
    if not isinstance(raw_value, str):
        return None
    normalized = normalize_text(raw_value)
    if not normalized:
        return None
    return INDUSTRY_LOOKUP.get(normalized)


def to_title_from_quote(quote: str) -> str:
    cleaned = quote.strip().strip('"').strip()
    if not cleaned:
        return "Use Case Signal"
    first_sentence = cleaned.split(".")[0].strip()
    if len(first_sentence) > 66:
        first_sentence = f"{first_sentence[:63].rstrip()}..."
    return first_sentence


def pct_change(previous: float, current: float) -> float:
    if previous <= 0:
        return 0.0
    return ((current - previous) / previous) * 100


def build_data(heatmatrix_path: Path, content_path: Path) -> dict[str, Any]:
    wb_heat = load_workbook(heatmatrix_path, data_only=True)
    wb_content = load_workbook(content_path, data_only=True)

    definitions: dict[str, str] = {}
    ws_definitions = wb_content["Definitions"]
    current_tech: str | None = None
    for row in range(1, ws_definitions.max_row + 1):
        cell_value = ws_definitions.cell(row, 1).value
        if not isinstance(cell_value, str):
            continue
        text = cell_value.strip()
        if normalize_text(text).startswith("2industries"):
            break
        if text.lower().startswith("definition:") and current_tech:
            definitions[current_tech] = text.replace("Definition:", "", 1).strip()
            continue
        maybe_tech = get_tech_key(text)
        if maybe_tech:
            current_tech = maybe_tech
            continue

    heat_scores: dict[str, list[dict[str, Any]]] = defaultdict(list)
    ws_heatmatrix = wb_heat["Heatmatrix"]
    heat_tech_keys = {
        column: get_tech_key(ws_heatmatrix.cell(2, column).value)
        for column in range(2, 7)
    }
    for row in range(3, 13):
        industry = get_industry(ws_heatmatrix.cell(row, 1).value)
        if not industry:
            continue
        industry_slug, industry_name = industry
        for column in range(2, 7):
            tech_key = heat_tech_keys.get(column)
            if not tech_key:
                continue
            score = parse_number(ws_heatmatrix.cell(row, column).value)
            heat_scores[tech_key].append(
                {
                    "industrySlug": industry_slug,
                    "industryName": industry_name,
                    "score": round(score, 2),
                }
            )

    radar_by_tech: dict[str, list[dict[str, Any]]] = defaultdict(list)
    ws_radar = wb_heat["Graph radar"]
    for row in range(4, ws_radar.max_row + 1):
        tech_key = get_tech_key(ws_radar.cell(row, 2).value)
        industry = get_industry(ws_radar.cell(row, 1).value)
        if not tech_key or not industry:
            continue
        industry_slug, industry_name = industry
        radar_by_tech[tech_key].append(
            {
                "industrySlug": industry_slug,
                "industryName": industry_name,
                "innovationIntensity": round(parse_number(ws_radar.cell(row, 3).value), 2),
                "innovationMomentum": round(parse_number(ws_radar.cell(row, 4).value), 2),
                "startupActivity": round(parse_number(ws_radar.cell(row, 5).value), 2),
                "marketValidation": round(parse_number(ws_radar.cell(row, 6).value), 2),
                "professionalsPerception": round(parse_number(ws_radar.cell(row, 7).value), 2),
            }
        )

    survey_by_tech: dict[str, dict[str, Any]] = {}
    ws_survey = wb_heat["Survey"]
    for spec in TECH_SPECS:
        survey_by_tech[spec.key] = {
            "impactScores": [],
            "timelineScores": [],
            "impactCounts": [0.0, 0.0, 0.0, 0.0, 0.0],
        }

    for row in range(2, 52):
        tech_key = get_tech_key(ws_survey.cell(row, 2).value)
        if not tech_key:
            continue
        item = survey_by_tech[tech_key]
        impact_score = ws_survey.cell(row, 12).value
        timeline_score = ws_survey.cell(row, 13).value
        if isinstance(impact_score, (int, float)):
            item["impactScores"].append(float(impact_score))
        if isinstance(timeline_score, (int, float)):
            item["timelineScores"].append(float(timeline_score))
        for offset in range(5):
            item["impactCounts"][offset] += parse_number(ws_survey.cell(row, 3 + offset).value)

    impact_ranking = []
    for spec in TECH_SPECS:
        counts = survey_by_tech[spec.key]["impactCounts"]
        total = sum(counts)
        if total <= 0:
            continue
        weighted_score = sum((index + 1) * counts[index] for index in range(5)) / total
        impact_ranking.append(
            {
                "technologyKey": spec.key,
                "technologySlug": spec.slug,
                "label": spec.name,
                "score": round(weighted_score, 2),
            }
        )
    impact_ranking.sort(key=lambda item: item["score"], reverse=True)

    full_table_aggregates: dict[str, dict[str, float]] = {}
    ws_full = wb_heat["FULL TABLE and WEIGHTS"]
    for spec in TECH_SPECS:
        full_table_aggregates[spec.key] = {
            "unicornFundingUsd": 0.0,
            "unicornCount": 0.0,
            "nativeUnicornCount": 0.0,
            "emergingUnicornCount": 0.0,
            "startupCount2025": 0.0,
            "patents2024": 0.0,
            "patents2025": 0.0,
            "scholar2024": 0.0,
            "scholar2025": 0.0,
        }

    for row in range(2, 52):
        tech_key = get_tech_key(ws_full.cell(row, 2).value)
        if not tech_key:
            continue
        aggregate = full_table_aggregates[tech_key]
        aggregate["unicornFundingUsd"] += parse_number(ws_full.cell(row, 3).value)
        aggregate["unicornCount"] += parse_number(ws_full.cell(row, 5).value)
        aggregate["nativeUnicornCount"] += parse_number(ws_full.cell(row, 7).value)
        aggregate["emergingUnicornCount"] += parse_number(ws_full.cell(row, 9).value)
        aggregate["startupCount2025"] += parse_number(ws_full.cell(row, 11).value)
        aggregate["patents2025"] += parse_number(ws_full.cell(row, 13).value)
        aggregate["scholar2025"] += parse_number(ws_full.cell(row, 15).value)
        aggregate["patents2024"] += parse_number(ws_full.cell(row, 17).value)
        aggregate["scholar2024"] += parse_number(ws_full.cell(row, 18).value)

    # Q9 use-case aggregation from Survey Results columns U:FN.
    ws_survey_results = wb_content["Survey Results"]
    q9_use_case_votes: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    for industry_name, start_col_letter, _end_col_letter in Q9_INDUSTRY_BLOCKS:
        block_start = column_index_from_string(start_col_letter)
        for tech_key, chunk_index in TECH_TO_USECASE_CHUNK_INDEX.items():
            chunk_start = block_start + chunk_index * 3
            chunk_end = chunk_start + 2
            for column in range(chunk_start, chunk_end + 1):
                label = clean_label(ws_survey_results.cell(2, column).value)
                if not label:
                    continue
                votes = 0
                for row in range(3, ws_survey_results.max_row + 1):
                    value = ws_survey_results.cell(row, column).value
                    if isinstance(value, str):
                        if value.strip():
                            votes += 1
                    elif value not in (None, ""):
                        votes += 1

                if votes <= 0:
                    continue

                existing = q9_use_case_votes[tech_key].setdefault(
                    label,
                    {"label": label, "votes": 0, "industries": set()},
                )
                existing["votes"] += votes
                existing["industries"].add(industry_name)

    q9_use_cases_by_tech: dict[str, list[dict[str, str]]] = {}
    for tech_key, vote_map in q9_use_case_votes.items():
        ranked = sorted(vote_map.values(), key=lambda item: (-item["votes"], item["label"]))[:6]
        q9_use_cases_by_tech[tech_key] = [
            {
                "title": item["label"],
                "description": f"Selected {item['votes']} times in Q9 across all industries.",
                "signal": f"Q9 aggregation · {len(item['industries'])} industry blocks",
            }
            for item in ranked
        ]

    chair_comment_placeholder = {
        "speaker": "Chair",
        "quote": "Chair's comment placeholder — final commentary will be inserted here.",
    }

    scatter_points: dict[str, dict[str, dict[str, Any]]] = defaultdict(dict)
    category_priority = {"Emerging": 1, "Unicorn": 2, "Native": 3}

    def ingest_scatter(
        ws_name: str,
        tech_col: int,
        org_col: int,
        year_col: int,
        funding_col: int,
        short_desc_col: int,
        full_desc_col: int,
        precision_col: int,
        category: str,
    ) -> None:
        ws = wb_content[ws_name]
        for row in range(2, ws.max_row + 1):
            tech_key = get_tech_key(ws.cell(row, tech_col).value)
            if not tech_key:
                continue
            organization = ws.cell(row, org_col).value
            if not isinstance(organization, str) or not organization.strip():
                continue
            founded_raw = ws.cell(row, year_col).value
            founded_year = parse_year(founded_raw)
            funding_usd = parse_number(ws.cell(row, funding_col).value)
            if not founded_year or funding_usd <= 0:
                continue
            description = clean_label(ws.cell(row, short_desc_col).value)
            if not description:
                description = clean_label(ws.cell(row, full_desc_col).value) or "No activity description available."
            founded_date = format_founded_date(founded_raw, ws.cell(row, precision_col).value)
            normalized_org = normalize_text(organization)

            candidate = {
                "name": organization.strip(),
                "foundedYear": founded_year,
                "foundedDate": founded_date,
                "fundingUsd": round(funding_usd, 2),
                "description": description,
                "category": category,
                "_priority": category_priority[category],
            }
            existing = scatter_points[tech_key].get(normalized_org)
            if not existing:
                scatter_points[tech_key][normalized_org] = candidate
                continue

            if candidate["_priority"] > existing["_priority"]:
                scatter_points[tech_key][normalized_org] = candidate
                continue

            if candidate["_priority"] == existing["_priority"] and candidate["fundingUsd"] > existing["fundingUsd"]:
                scatter_points[tech_key][normalized_org] = candidate

    ingest_scatter(
        "Native Unicorns",
        tech_col=2,
        org_col=3,
        year_col=10,
        funding_col=12,
        short_desc_col=7,
        full_desc_col=8,
        precision_col=11,
        category="Native",
    )
    ingest_scatter(
        "Unicorns - TECH",
        tech_col=1,
        org_col=2,
        year_col=8,
        funding_col=10,
        short_desc_col=5,
        full_desc_col=6,
        precision_col=9,
        category="Unicorn",
    )
    ingest_scatter(
        "Emerging Unicorns - TECH",
        tech_col=1,
        org_col=2,
        year_col=9,
        funding_col=11,
        short_desc_col=6,
        full_desc_col=7,
        precision_col=10,
        category="Emerging",
    )

    patents_by_tech: dict[str, dict[int, float]] = defaultdict(dict)
    scholar_by_tech: dict[str, dict[int, float]] = defaultdict(dict)

    ws_patents_graph = wb_content["Patents evol Tech GRAPH"]
    tech_columns_patents = {
        column: get_tech_key(ws_patents_graph.cell(2, column).value) for column in range(2, 7)
    }
    for row in range(3, ws_patents_graph.max_row + 1):
        year = parse_year(ws_patents_graph.cell(row, 1).value)
        if not year:
            continue
        for column in range(2, 7):
            tech_key = tech_columns_patents.get(column)
            if not tech_key:
                continue
            patents_by_tech[tech_key][year] = round(parse_number(ws_patents_graph.cell(row, column).value), 2)

    ws_scholar_graph = wb_content["Scholar evol TECH graph"]
    tech_columns_scholar = {
        column: get_tech_key(ws_scholar_graph.cell(2, column).value) for column in range(2, 7)
    }
    for row in range(3, ws_scholar_graph.max_row + 1):
        year = parse_year(ws_scholar_graph.cell(row, 1).value)
        if not year:
            continue
        for column in range(2, 7):
            tech_key = tech_columns_scholar.get(column)
            if not tech_key:
                continue
            scholar_by_tech[tech_key][year] = round(parse_number(ws_scholar_graph.cell(row, column).value), 2)

    patents_series = []
    scholar_series = []
    for year in RESEARCH_POINT_YEARS:
        patents_entry: dict[str, Any] = {"year": str(year)}
        scholar_entry: dict[str, Any] = {"year": str(year)}
        for spec in TECH_SPECS:
            patents_entry[spec.slug] = patents_by_tech[spec.key].get(year, 0)
            scholar_entry[spec.slug] = scholar_by_tech[spec.key].get(year, 0)
        patents_series.append(patents_entry)
        scholar_series.append(scholar_entry)

    applicants_by_tech: dict[str, list[dict[str, Any]]] = defaultdict(list)
    ws_top_applicants = wb_content["Top Applicants 2025 - TECH"]
    for row in range(2, ws_top_applicants.max_row + 1):
        tech_key = get_tech_key(ws_top_applicants.cell(row, 1).value)
        applicant_name = ws_top_applicants.cell(row, 2).value
        if not tech_key or not isinstance(applicant_name, str) or not applicant_name.strip():
            continue
        status_raw = ws_top_applicants.cell(row, 4).value
        status = "none"
        if isinstance(status_raw, str):
            normalized = normalize_text(status_raw)
            if "newchallenger" in normalized:
                status = "new_challenger"
            elif "significantclimber" in normalized:
                status = "significant_climber"
            elif "significantdrop" in normalized:
                status = "significant_drop"
        applicants_by_tech[tech_key].append(
            {
                "name": applicant_name.strip(),
                "documentCount": int(parse_number(ws_top_applicants.cell(row, 3).value)),
                "status": status,
            }
        )

    for tech_key in applicants_by_tech:
        applicants_by_tech[tech_key].sort(key=lambda item: item["documentCount"], reverse=True)

    notes_by_tech: dict[str, str] = {}
    ws_notes = wb_content["Top Applicants 2025 note - TECH"]
    for row in range(2, ws_notes.max_row + 1):
        tech_key = get_tech_key(ws_notes.cell(row, 1).value)
        note = ws_notes.cell(row, 2).value
        if tech_key and isinstance(note, str) and note.strip():
            notes_by_tech[tech_key] = note.strip().replace("Note:", "").strip()

    technology_pages: list[dict[str, Any]] = []
    for spec in TECH_SPECS:
        industry_heat = heat_scores.get(spec.key, [])
        industry_radar_lookup = {entry["industrySlug"]: entry for entry in radar_by_tech.get(spec.key, [])}
        merged_industry_cards = []
        for heat_entry in industry_heat:
            radar_entry = industry_radar_lookup.get(heat_entry["industrySlug"])
            if not radar_entry:
                continue
            merged_industry_cards.append(
                {
                    "industrySlug": heat_entry["industrySlug"],
                    "industryName": heat_entry["industryName"],
                    "heatScore": heat_entry["score"],
                    "radar": {
                        "innovationIntensity": radar_entry["innovationIntensity"],
                        "innovationMomentum": radar_entry["innovationMomentum"],
                        "startupActivity": radar_entry["startupActivity"],
                        "marketValidation": radar_entry["marketValidation"],
                        "professionalsPerception": radar_entry["professionalsPerception"],
                    },
                }
            )

        avg_heat = round(mean(item["score"] for item in industry_heat), 1) if industry_heat else 0.0
        top_sector = max(industry_heat, key=lambda item: item["score"])["industryName"] if industry_heat else "n/a"
        cells_above_60 = len([item for item in industry_heat if item["score"] >= 60])

        impact_counts = survey_by_tech[spec.key]["impactCounts"]
        total_impact = sum(impact_counts)
        labels = [
            "No impact",
            "Minor impact",
            "Moderate impact",
            "Significant impact",
            "Very significant impact",
        ]
        impact_distribution = []
        for index, label in enumerate(labels):
            count = int(impact_counts[index])
            pct = round((count / total_impact) * 100, 1) if total_impact else 0.0
            impact_distribution.append({"label": label, "count": count, "percentage": pct})

        market = full_table_aggregates[spec.key]
        patents_delta = pct_change(market["patents2024"], market["patents2025"])
        scholar_delta = pct_change(market["scholar2024"], market["scholar2025"])

        tech_use_cases = q9_use_cases_by_tech.get(spec.key, [])[:6]
        points = []
        for point in scatter_points.get(spec.key, {}).values():
            cleaned = dict(point)
            cleaned.pop("_priority", None)
            points.append(cleaned)
        points.sort(key=lambda item: item["fundingUsd"], reverse=True)

        technology_pages.append(
            {
                "id": spec.key,
                "slug": spec.slug,
                "name": spec.name,
                "shortName": spec.short_name,
                "themeColor": spec.color,
                "definition": definitions.get(spec.key, ""),
                "summary": f"{spec.name} records an average heat score of {avg_heat}/100 across the ten sectors in the 2026 matrix.",
                "overview": {
                    "summaryMetrics": {
                        "averageHeatScore": avg_heat,
                        "topExposureSector": top_sector,
                        "cellsAbove60": cells_above_60,
                    },
                    "industryRadarProfiles": merged_industry_cards,
                    "chairComment": chair_comment_placeholder,
                },
                "professionalsPerception": {
                    "impactRanking": impact_ranking,
                    "impactDistribution": impact_distribution,
                    "topUseCases": tech_use_cases,
                },
                "marketValidation": {
                    "totals": {
                        "unicornFundingUsd": round(market["unicornFundingUsd"], 2),
                        "unicornCount": int(market["unicornCount"]),
                        "nativeUnicornCount": int(market["nativeUnicornCount"]),
                        "emergingUnicornCount": int(market["emergingUnicornCount"]),
                        "startupCount2025": int(market["startupCount2025"]),
                    },
                    "scatterPoints": points,
                },
                "researchInnovation": {
                    "patentsDeltaPct": round(patents_delta, 1),
                    "scholarDeltaPct": round(scholar_delta, 1),
                    "topApplicants2025": applicants_by_tech.get(spec.key, [])[:16],
                    "topApplicantsNote": notes_by_tech.get(spec.key, ""),
                },
            }
        )

    return {
        "generatedAt": datetime.now().isoformat(),
        "source": {
            "heatmatrixWorkbook": str(heatmatrix_path),
            "contentWorkbook": str(content_path),
        },
        "technologyOrder": [
            {
                "id": spec.key,
                "slug": spec.slug,
                "name": spec.name,
                "color": spec.color,
            }
            for spec in TECH_SPECS
        ],
        "researchSeries": {
            "patents": patents_series,
            "scholarWork": scholar_series,
        },
        "technologies": technology_pages,
    }


def main() -> None:
    project_root = Path(__file__).resolve().parents[1]
    parser = argparse.ArgumentParser(description="Build technology page data JSON from source workbooks.")
    parser.add_argument(
        "--heatmatrix",
        type=Path,
        default=Path.home() / "Downloads" / "Heatmatrix 2026 Deciles-2.xlsx",
        help="Path to Heatmatrix workbook",
    )
    parser.add_argument(
        "--content",
        type=Path,
        default=Path.home() / "Downloads" / "Matrix content table-5.xlsx",
        help="Path to matrix content workbook",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=project_root / "src" / "data" / "raw" / "technologyPages.json",
        help="Output JSON path",
    )
    args = parser.parse_args()

    payload = build_data(args.heatmatrix, args.content)
    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8") as file:
        json.dump(payload, file, indent=2, ensure_ascii=False)
        file.write("\n")
    print(f"Generated {args.output}")


if __name__ == "__main__":
    main()
