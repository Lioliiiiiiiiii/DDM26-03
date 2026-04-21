"""
Extract and normalize the tech-overview data from the two Excel source files
into a single JSON file consumed by the redesigned technology pages.

Inputs (expected in ~/Downloads):
  - Heatmatrix 2026 Deciles-2.xlsx
  - Matrix content table-5.xlsx

Output:
  - assets/data/tech-data.json (relative to repo root)
"""
import json
import os
from pathlib import Path
import openpyxl
from datetime import datetime, date

REPO = Path(__file__).resolve().parents[1]
DOWNLOADS = Path.home() / "Downloads"
HEATMATRIX_XLSX = DOWNLOADS / "Heatmatrix 2026 Deciles-2.xlsx"
CONTENT_XLSX = DOWNLOADS / "Matrix content table-5.xlsx"
OUT = REPO / "assets" / "data" / "tech-data.json"

# Normalized technology mapping (as specified by the spec).
TECHS = {
    "descriptive-ai": {
        "pageTitle": "Descriptive AI",
        "definitionSheetLabel": "Descriptive AI (before was called traditional AI)",
        "definitionText": "Descriptive AI encompasses traditional AI approaches that analyze and interpret existing data to identify patterns, make predictions, and derive insights. This technology forms the foundation of practical AI applications in industry through machine learning, statistical analysis, and pattern recognition.",
        "heatmatrixColumn": "Descriptive AI",
        "radarTech": "Descriptive AI",
        "patentScholarTech": "Descriptive AI",
        "topApplicantsTech": "Descriptive AI",
        "unicornTech": "Descriptive AI",
        "surveyRankHeader": "Descriptive AI",
        "surveyImpactHeader": "Descriptive AI",
        "useCaseChunkIndex": 0,
    },
    "agentic-genai": {
        "pageTitle": "Agentic & GenAI",
        "definitionSheetLabel": "Generative AI & Agentic AI",
        "definitionText": "Generative AI and Agentic AI refers to AI systems that can create new content (text, images, code, synthetic data) and autonomously plan and execute multi-step actions to achieve goals. This technology leverages large language models and neural networks to generate novel outputs based on patterns learned from training data and complete tasks with reduced human supervision.",
        "heatmatrixColumn": "Agentic & Gen AI",
        "radarTech": "Agentic&GenAI",
        "patentScholarTech": "Agentic & Gen AI",
        "topApplicantsTech": "Agentic & GenAI",
        "unicornTech": "Agentic & GenAI",
        "surveyRankHeader": "Generative AI & Agentic AI",
        "surveyImpactHeader": "Generative AI & Agentic AI",
        "useCaseChunkIndex": 1,
    },
    "blockchain-decentralized-systems": {
        "pageTitle": "Blockchain & Decentralized Systems",
        "definitionSheetLabel": "Blockchain & Decentralized Systems",
        "definitionText": "Blockchain technology enables secure, decentralized record-keeping and transactions without requiring central authority. It creates transparent, immutable records across distributed networks, supporting applications from cryptocurrency to supply chain tracking.",
        "heatmatrixColumn": "Blockchain",
        "radarTech": "Blockchain",
        "patentScholarTech": "Blockchain",
        "topApplicantsTech": "Blockchain",
        "unicornTech": "Blockchain",
        "surveyRankHeader": "Blockchain & Decentralized Systems",
        "surveyImpactHeader": "Blockchain & Decentralized Systems",
        "useCaseChunkIndex": 2,
    },
    "robotics-automation": {
        "pageTitle": "Robotics & Automation",
        "definitionSheetLabel": "Robotics",
        "definitionText": "Robotics and Automation encompasses the design and deployment of physical and virtual systems that can perform tasks automatically with minimal human intervention. This includes industrial robots, autonomous systems, and process automation technologies.",
        "heatmatrixColumn": "Robotics & Automation",
        "radarTech": "Robotics",
        "patentScholarTech": "Robotics",
        "topApplicantsTech": "Robotics",
        "unicornTech": "Robotics",
        "surveyRankHeader": "Robotics",
        "surveyImpactHeader": "Robotics",
        "useCaseChunkIndex": 3,
    },
    "quantum-computing": {
        "pageTitle": "Quantum Computing",
        "definitionSheetLabel": "Quantum Computing",
        "definitionText": "Quantum Computing harnesses quantum mechanical phenomena like superposition and entanglement to perform computations. This technology enables exponentially faster calculations for specific problems, particularly in cryptography, material science, and complex system optimization.",
        "heatmatrixColumn": "Quantum Computing",
        "radarTech": "Quantum",
        "patentScholarTech": "Quantum",
        "topApplicantsTech": "Quantum",
        "unicornTech": "Quantum",
        "surveyRankHeader": "Quantum Computing",
        "surveyImpactHeader": "Quantum Computing",
        "useCaseChunkIndex": 4,
    },
}

# Display industry order for the first analysis card and for the heatmatrix rows.
INDUSTRY_ORDER = [
    "Energy",
    "Materials",
    "Industrials",
    "Consumer Goods",
    "Health Care",
    "Financial services",
    "Information technology",
    "Communication & creative services",
    "Real Estate",
    "Automotive & transport",
]

# Radar sheet uses compact industry labels without spaces.
RADAR_INDUSTRY_ALIAS = {
    "Energy": "Energy",
    "Materials": "Materials",
    "Industrials": "Industrials",
    "Consumer Goods": "Consumer Goods",
    "Health Care": "HealthCare",
    "Financial services": "Financialsservices",
    "Information technology": "Informationtechnology",
    "Communication & creative services": "Communicationservices(&creative)",
    "Real Estate": "RealEstate",
    "Automotive & transport": "Automotive&transport",
}

# Survey raw-data industry labels (first column) → display label.
SURVEY_INDUSTRY_PREFIX = {
    "Energy": "Energy",
    "Industrials": "Industrials",
    "Materials": "Materials",
    "Consumer Goods": "Consumer Goods",
    "Health Care": "Health Care",
    "Financial services": "Financial services",
    "Information technology": "Information technology",
    "Real Estate": "Real Estate",
    "Communication services (& creative)": "Communication & creative services",
    "Automotive & transport": "Automotive & transport",
}

USE_CASE_INDUSTRY_ORDER = [
    ("Energy", 21),
    ("Industrials", 36),
    ("Materials", 51),
    ("Consumer Goods", 66),
    ("Health Care", 81),
    ("Financial services", 96),
    ("Information technology", 111),
    ("Real Estate", 126),
    ("Communication & creative services", 141),
    ("Automotive & transport", 156),
]

RADAR_AXES = [
    "Innovation Intensity",
    "innovation Momentum",
    "Startup Activity",
    "Market Validation",
    "Professional's Perception",
]


def load_workbook(path):
    return openpyxl.load_workbook(path, data_only=True)


def extract_heatmatrix(wb):
    ws = wb["Heatmatrix"]
    header_row = [c.value for c in ws[2]]
    tech_cols = {}
    for idx, val in enumerate(header_row):
        if val in ("Descriptive AI", "Agentic & Gen AI", "Blockchain", "Robotics & Automation", "Quantum Computing"):
            tech_cols[val] = idx
    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        sector = r[0]
        if not sector:
            continue
        if sector == "Total must = 10":
            continue
        entry = {"sector": sector}
        for tech_name, idx in tech_cols.items():
            v = r[idx]
            entry[tech_name] = float(v) if isinstance(v, (int, float)) else None
        rows.append(entry)
    return {"columns": list(tech_cols.keys()), "rows": rows}


def extract_radar(wb):
    ws = wb["Graph radar"]
    out = {}
    for r in ws.iter_rows(min_row=4, values_only=True):
        industry = r[0]
        tech = r[1]
        if not industry or not tech:
            continue
        vals = list(r[2:7])
        out.setdefault(industry, {})[tech] = [
            float(v) if isinstance(v, (int, float)) else 0.0 for v in vals
        ]
    return out


def extract_patents_scholar(wb, sheet):
    ws = wb[sheet]
    header = [c.value for c in ws[2]]
    years = []
    data = {}
    for row in ws.iter_rows(min_row=3, values_only=True):
        tech = row[0]
        if not tech:
            continue
        if not years:
            years = [int(h) for h in header[1:] if isinstance(h, (int, float))]
        series = [float(v) if isinstance(v, (int, float)) else None for v in row[1:]]
        data[tech] = series
    return {"years": years, "series": data}


def extract_top_applicants(wb):
    ws = wb["Top Applicants 2025 - TECH"]
    out = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        tech, name, count, tag = row[:4]
        if not tech or not name:
            continue
        out.setdefault(tech, []).append({
            "name": str(name).strip(),
            "count": float(count) if isinstance(count, (int, float)) else 0,
            "tag": tag or "",
        })
    # Notes
    notes = {}
    ws = wb["Top Applicants 2025 note - TECH"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        tech, note = row[0], row[1]
        if tech:
            notes[tech] = note or ""
    return {"byTech": out, "notes": notes}


def normalize_date(v):
    if isinstance(v, (datetime, date)):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, (int, float)):
        return f"{int(v)}-01-01"
    if isinstance(v, str):
        return v
    return None


def extract_companies(wb):
    """Collect companies for Native / Unicorn / Emerging categories per tech.
    Dedupe per tech by organization name using priority: Native > Unicorn > Emerging.
    """
    out = {}  # tech -> list[dict]

    def push(tech, entry):
        if not tech:
            return
        out.setdefault(tech, []).append(entry)

    # Native Unicorns
    ws = wb["Native Unicorns"]
    # header at row 1; columns: Industry, Technology, Organization Name, URL, Industries,
    # Industry Groups, Description, Full Description, Hub Tags, Founded Date, Precision,
    # Total Funding Amount (in USD)
    header = [c.value for c in ws[1]]
    for row in ws.iter_rows(min_row=2, values_only=True):
        tech = row[1]
        name = row[2]
        if not name:
            continue
        push(tech, {
            "category": "Native",
            "name": str(name).strip(),
            "description": row[6] or "",
            "fullDescription": row[7] or "",
            "founded": normalize_date(row[9]),
            "funding": float(row[11]) if isinstance(row[11], (int, float)) else None,
        })

    # Unicorns - TECH
    ws = wb["Unicorns - TECH"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        tech = row[0]
        name = row[1]
        if not name:
            continue
        push(tech, {
            "category": "Unicorn",
            "name": str(name).strip(),
            "description": row[4] or "",
            "fullDescription": row[5] or "",
            "founded": normalize_date(row[7]),
            "funding": float(row[9]) if isinstance(row[9], (int, float)) else None,
        })

    # Emerging Unicorns - TECH
    ws = wb["Emerging Unicorns - TECH"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        tech = row[0]
        name = row[1]
        if not name:
            continue
        push(tech, {
            "category": "Emerging",
            "name": str(name).strip(),
            "description": row[5] or "",
            "fullDescription": row[6] or "",
            "founded": normalize_date(row[8]),
            "funding": float(row[10]) if isinstance(row[10], (int, float)) else None,
        })

    # Deduplicate per tech with priority Native > Unicorn > Emerging.
    priority = {"Native": 0, "Unicorn": 1, "Emerging": 2}
    deduped = {}
    for tech, items in out.items():
        best = {}
        for it in items:
            key = it["name"].lower()
            if key not in best or priority[it["category"]] < priority[best[key]["category"]]:
                best[key] = it
        deduped[tech] = list(best.values())
    return deduped


def extract_survey(wb):
    """Extract ranking mean, per-tech impact distribution, and use cases from Survey Results."""
    ws = wb["Survey Results"]
    # row 2 (1-indexed) has tech labels. We iterate data rows from row 3.
    row2 = [c.value for c in ws[2]]

    # F:J (6-10) Q6 ranks; K:O (11-15) Q7 impact.
    rank_cols = {
        "Descriptive AI": 6 - 1,
        "Generative AI & Agentic AI": 7 - 1,
        "Blockchain & Decentralized Systems": 8 - 1,
        "Robotics": 9 - 1,
        "Quantum Computing": 10 - 1,
    }
    impact_cols = {
        "Descriptive AI": 11 - 1,
        "Generative AI & Agentic AI": 12 - 1,
        "Blockchain & Decentralized Systems": 13 - 1,
        "Robotics": 14 - 1,
        "Quantum Computing": 15 - 1,
    }

    BUCKETS = [
        "No impact",
        "Minor impact",
        "Moderate impact",
        "Significant impact",
        "Very significant impact",
    ]

    rank_sum = {k: 0.0 for k in rank_cols}
    rank_n = {k: 0 for k in rank_cols}
    impact_counts = {k: {b: 0 for b in BUCKETS} for k in impact_cols}

    # Q9 use cases: 10 industry blocks × 15 cols starting col index 20 (U = 21).
    # Row 2 has the use-case labels.
    use_case_labels = []
    for industry, start_col in USE_CASE_INDUSTRY_ORDER:
        # start_col is 1-indexed (U = 21). Convert to 0-indexed.
        base = start_col - 1
        block = []
        for i in range(15):
            block.append(row2[base + i])
        use_case_labels.append((industry, start_col, block))

    # Counts: industry x col-offset
    use_case_counts = {
        industry: [0] * 15 for industry, _, _ in use_case_labels
    }

    for ri, row in enumerate(ws.iter_rows(min_row=3, values_only=True), start=3):
        # Rank columns (numbers 1-5)
        for tech, idx in rank_cols.items():
            v = row[idx]
            if isinstance(v, (int, float)):
                rank_sum[tech] += float(v)
                rank_n[tech] += 1
        # Impact columns (bucket strings)
        for tech, idx in impact_cols.items():
            v = row[idx]
            if isinstance(v, str) and v in impact_counts[tech]:
                impact_counts[tech][v] += 1
        # Use case columns: any truthy value counts as a selection.
        for industry, start_col, _ in use_case_labels:
            base = start_col - 1
            for i in range(15):
                v = row[base + i]
                if v not in (None, "", 0, 0.0):
                    use_case_counts[industry][i] += 1

    ranking = []
    for tech in rank_cols:
        mean = (rank_sum[tech] / rank_n[tech]) if rank_n[tech] else None
        ranking.append({"tech": tech, "mean": mean, "n": rank_n[tech]})
    ranking.sort(key=lambda x: (x["mean"] is None, x["mean"] or 0))

    # Flatten use-case labels so the front-end can map chunk index → use case.
    return {
        "ranking": ranking,
        "impactBuckets": BUCKETS,
        "impactDistribution": impact_counts,
        "useCases": [
            {"industry": ind, "labels": labels, "counts": use_case_counts[ind]}
            for ind, _, labels in use_case_labels
        ],
    }


def main():
    hm_wb = load_workbook(HEATMATRIX_XLSX)
    c_wb = load_workbook(CONTENT_XLSX)

    data = {
        "techs": TECHS,
        "industryOrder": INDUSTRY_ORDER,
        "radarIndustryAlias": RADAR_INDUSTRY_ALIAS,
        "radarAxes": RADAR_AXES,
        "heatmatrix": extract_heatmatrix(hm_wb),
        "radar": extract_radar(hm_wb),
        "patents": extract_patents_scholar(c_wb, "Patents Evolution - TECH"),
        "scholar": extract_patents_scholar(c_wb, "Scholar evolution - TECH"),
        "topApplicants": extract_top_applicants(c_wb),
        "companies": extract_companies(c_wb),
        "survey": extract_survey(c_wb),
    }

    OUT.parent.mkdir(parents=True, exist_ok=True)
    with open(OUT, "w") as f:
        json.dump(data, f, separators=(",", ":"), ensure_ascii=False)
    print("Wrote", OUT, "bytes:", OUT.stat().st_size)

    # Print sanity checks
    print("\nMean rank (lower = stronger impact):")
    for r in data["survey"]["ranking"]:
        print(f"  {r['tech']}: {r['mean']:.2f}  (n={r['n']})")

    print("\nPatents 2024→2025 change:")
    for tech, series in data["patents"]["series"].items():
        if series[-2] and series[-1]:
            pct = (series[-1] - series[-2]) / series[-2] * 100
            print(f"  {tech}: {pct:+.1f}%")

    print("\nScholar 2024→2025 change:")
    for tech, series in data["scholar"]["series"].items():
        if series[-2] and series[-1]:
            pct = (series[-1] - series[-2]) / series[-2] * 100
            print(f"  {tech}: {pct:+.1f}%")


if __name__ == "__main__":
    main()
